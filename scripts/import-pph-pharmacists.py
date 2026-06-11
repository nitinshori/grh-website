#!/usr/bin/env python3
"""
Bulk-import PPH pharmacists from Jane's password-protected spreadsheet.

Reads the locked .xlsx, parses pharmacist rows, generates a temp
password + bcrypt hash for each, and writes two files:
  - pph-credentials.csv : GPHC, name, branch, temp password
                          (give this to Jane to distribute)
  - pph-inserts.sql     : INSERT statements you can run against Neon

Dependencies (one-off):
    pip install msoffcrypto-tool openpyxl bcrypt --break-system-packages

Usage:
    python3 scripts/import-pph-pharmacists.py \
        "/path/to/GRH PGD Branch Access Information.xlsx" \
        --password GrhPPH \
        --pph-pharmacy-id <PPH_PHARMACY_UUID>

Then review pph-inserts.sql and apply:
    psql "$DATABASE_URL" -f scripts/output/pph-inserts.sql
"""

import argparse, csv, io, os, secrets, sys, uuid
from pathlib import Path

try:
    import msoffcrypto
    from openpyxl import load_workbook
    import bcrypt
except ImportError as e:
    sys.exit(
        f"Missing dependency ({e.name}). Install with:\n"
        "  pip install msoffcrypto-tool openpyxl bcrypt --break-system-packages\n"
    )


# ── arg parsing ─────────────────────────────────────────────────

parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
parser.add_argument("xlsx_path", help="Path to the locked .xlsx file")
parser.add_argument("--password", required=True, help="Password for the xlsx")
parser.add_argument("--pph-pharmacy-id", required=True,
                    help="UUID of the Pharmacy Plus Health row in the pharmacies table")
parser.add_argument("--output-dir", default="scripts/output",
                    help="Where to write the CSV and SQL files (default: scripts/output)")
args = parser.parse_args()

xlsx_path = Path(args.xlsx_path)
if not xlsx_path.exists():
    sys.exit(f"File not found: {xlsx_path}")

output_dir = Path(args.output_dir)
output_dir.mkdir(parents=True, exist_ok=True)


# ── 1. decrypt + open the workbook ──────────────────────────────

with open(xlsx_path, "rb") as f:
    file = msoffcrypto.OfficeFile(f)
    file.load_key(password=args.password)
    decrypted = io.BytesIO()
    file.decrypt(decrypted)

wb = load_workbook(decrypted, data_only=True)
ws = wb[wb.sheetnames[0]]


# ── 2. dedupe pharmacists across branches ───────────────────────

def clean_name(s):
    if not s:
        return ""
    s = str(s).strip()
    if s.lower() in ("n.a", "n/a", "none"):
        return ""
    # Strip the annotations Jane added inline
    for tag in ("(MAT LEAVE)", "(LOCUM COVERING MAT LEAVE)", "(MAT LEAVE)"):
        s = s.replace(tag, "")
    return s.strip()


def clean_gphc(s):
    if not s:
        return ""
    s = str(s).strip()
    if s.lower() in ("n.a", "n/a", "none"):
        return ""
    return s


pharmacists = {}  # gphc -> { name, branches[] }
header = [str(c.value or "").strip() for c in ws[1]]

# Column indices (1-based, then we'll switch to 0-based)
COL_BRANCH = 0
COL_MAIN_PHARM = 6
COL_MAIN_GPHC = 7
COL_LOCUM_PHARM = 8
COL_LOCUM_GPHC = 9

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[COL_BRANCH]:
        continue
    branch = str(row[COL_BRANCH]).strip()
    for pcol, gcol in ((COL_MAIN_PHARM, COL_MAIN_GPHC), (COL_LOCUM_PHARM, COL_LOCUM_GPHC)):
        name = clean_name(row[pcol] if pcol < len(row) else "")
        gphc = clean_gphc(row[gcol] if gcol < len(row) else "")
        if not name or not gphc:
            continue
        if gphc not in pharmacists:
            pharmacists[gphc] = {"name": name, "branches": []}
        pharmacists[gphc]["branches"].append(branch)

print(f"Parsed {len(pharmacists)} unique pharmacists from {xlsx_path.name}")


# ── 3. generate temp passwords + bcrypt hashes ──────────────────

# Unambiguous alphabet (no 0/O, no 1/l/I). 12-char temp password is
# enough entropy for a one-time-use credential that's force-rotated
# on first login.
ALPHA = "ABCDEFGHJKMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"

def temp_password():
    return "".join(secrets.choice(ALPHA) for _ in range(12))


credentials = []
for gphc, p in pharmacists.items():
    name = p["name"]
    parts = name.split()
    first_name = parts[0] if parts else ""
    last_name = " ".join(parts[1:]) if len(parts) > 1 else "(unknown)"
    tmp = temp_password()
    pw_hash = bcrypt.hashpw(tmp.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")
    credentials.append({
        "gphc": gphc,
        "name": name,
        "first_name": first_name,
        "last_name": last_name,
        "branches": " / ".join(p["branches"]),
        "temp_password": tmp,
        "password_hash": pw_hash,
    })

print(f"Generated temp passwords + bcrypt hashes for {len(credentials)} accounts")


# ── 4. write credentials CSV (for Jane) ─────────────────────────

csv_path = output_dir / "pph-credentials.csv"
with open(csv_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["GPHC (username)", "Name", "Branch(es)", "Temporary password"])
    for c in credentials:
        w.writerow([c["gphc"], c["name"], c["branches"], c["temp_password"]])

print(f"Wrote credentials CSV: {csv_path}")
print("  -> Send this to Jane. She distributes to each pharmacist.")


# ── 5. write SQL INSERT file (for Nitin to psql) ────────────────

sql_path = output_dir / "pph-inserts.sql"


def sql_quote(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


with open(sql_path, "w", encoding="utf-8") as f:
    f.write("-- PPH pharmacist bulk import\n")
    f.write(f"-- Source: {xlsx_path}\n")
    f.write(f"-- Pharmacy ID: {args.pph_pharmacy_id}\n")
    f.write(f"-- Total accounts: {len(credentials)}\n\n")
    f.write("BEGIN;\n\n")
    for c in credentials:
        synthetic_email = f"gphc-{c['gphc']}@pph.grh.internal"
        f.write(
            "INSERT INTO users ("
            "email, password_hash, first_name, last_name, role, pharmacy_id, "
            "is_active, auth_source, username, must_change_password"
            ") VALUES ("
            f"{sql_quote(synthetic_email)}, "
            f"{sql_quote(c['password_hash'])}, "
            f"{sql_quote(c['first_name'])}, "
            f"{sql_quote(c['last_name'])}, "
            f"'pharmacist', "
            f"'{args.pph_pharmacy_id}', "
            f"true, "
            f"'pph', "
            f"{sql_quote(c['gphc'])}, "
            f"true"
            ") ON CONFLICT (email) DO NOTHING;\n"
        )
    f.write("\nCOMMIT;\n")

print(f"Wrote SQL inserts:    {sql_path}")
print(f"  -> Apply with: psql \"$DATABASE_URL\" -f {sql_path}")
print("")
print("First login flow for each pharmacist:")
print("  1. Visit getrealhealthpgd.co.uk/login")
print("  2. Enter their GPHC number + temp password from the CSV")
print("  3. Forced to /change-password — sets a real password")
print("  4. Redirected to dashboard")
